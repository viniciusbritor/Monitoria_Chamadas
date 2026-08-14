import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _extract_call_fields(c: dict) -> dict:
    raw = {}
    if c.get("raw_evaluation"):
        if isinstance(c["raw_evaluation"], dict):
            raw = c["raw_evaluation"]
        elif isinstance(c["raw_evaluation"], str):
            try:
                raw = json.loads(c["raw_evaluation"])
            except Exception:
                raw = {}
                
    atendente = (
        raw.get("nome_atendente") 
        or c.get("nome_atendente") 
        or c.get("atendente") 
        or c.get("operador") 
        or "Não Identificado"
    )
    
    motivo = (
        raw.get("motivo_contato") 
        or c.get("motivo_contato") 
        or c.get("motivo") 
        or "-"
    )
    
    classif = (
        raw.get("classificacao_motivo") 
        or c.get("classificacao_motivo") 
        or c.get("classificacao") 
        or "-"
    )
    
    humor_cli = raw.get("humor_cliente") or c.get("humor_cliente")
    if not humor_cli and (raw.get("sentimentos_cliente") or c.get("sentimentos_cliente")):
        s_cli = raw.get("sentimentos_cliente") or c.get("sentimentos_cliente")
        if isinstance(s_cli, str):
            try:
                s_cli = json.loads(s_cli)
            except Exception:
                s_cli = []
        if isinstance(s_cli, list) and len(s_cli) > 0 and isinstance(s_cli[0], dict):
            humor_cli = s_cli[0].get("sentimento", "-")
    humor_cli = humor_cli or "-"
    
    humor_op = (
        raw.get("humor_expert") 
        or raw.get("humor_operador") 
        or raw.get("humor_atendente") 
        or c.get("humor_expert") 
        or c.get("humor_operador") 
        or c.get("humor_atendente")
    )
    if not humor_op and (raw.get("sentimentos_operador") or c.get("sentimentos_operador")):
        s_op = raw.get("sentimentos_operador") or c.get("sentimentos_operador")
        if isinstance(s_op, str):
            try:
                s_op = json.loads(s_op)
            except Exception:
                s_op = []
        if isinstance(s_op, list) and len(s_op) > 0 and isinstance(s_op[0], dict):
            humor_op = s_op[0].get("sentimento", "-")
    humor_op = humor_op or "-"
    
    fases = raw.get("fases") or c.get("fases") or {}
    fase1 = fases.get("apresentacao") or fases.get("fase_1") or fases.get("inicio") or {}
    fase2 = fases.get("resolucao") or fases.get("fase_2") or fases.get("desenvolvimento") or {}
    fase3 = fases.get("fechamento") or fases.get("fase_3") or fases.get("conclusao") or {}
    
    fase1_qa = fase1.get("nota_qa") if fase1.get("nota_qa") is not None else "-"
    fase1_nps = fase1.get("nota_nps") if fase1.get("nota_nps") is not None else "-"
    fase1_analise = fase1.get("analise") or "-"
    
    fase2_qa = fase2.get("nota_qa") if fase2.get("nota_qa") is not None else "-"
    fase2_nps = fase2.get("nota_nps") if fase2.get("nota_nps") is not None else "-"
    fase2_analise = fase2.get("analise") or "-"
    
    fase3_qa = fase3.get("nota_qa") if fase3.get("nota_qa") is not None else "-"
    fase3_nps = fase3.get("nota_nps") if fase3.get("nota_nps") is not None else "-"
    fase3_analise = fase3.get("analise") or "-"
    
    p_pos_list = raw.get("pontos_positivos") or c.get("pontos_positivos") or []
    if isinstance(p_pos_list, list):
        p_pos = "\n".join(f"• {p}" for p in p_pos_list) if p_pos_list else "-"
    else:
        p_pos = str(p_pos_list)
        
    p_neg_list = raw.get("pontos_melhoria") or c.get("pontos_melhoria") or []
    if isinstance(p_neg_list, list):
        p_neg = "\n".join(f"• {p}" for p in p_neg_list) if p_neg_list else "-"
    else:
        p_neg = str(p_neg_list)
        
    rec_trein = (
        raw.get("recomendacao_treinamento") 
        or c.get("recomendacao_treinamento") 
        or "-"
    )
    
    nota_qa = raw.get("nota_geral") or c.get("nota") or c.get("nota_geral") or 0
    nota_op = raw.get("nota_qualidade_operador") or c.get("nota_qualidade_operador") or "-"
    nps_cli = raw.get("nota_sentimento_cliente") or c.get("nota_sentimento_cliente") or c.get("nps") or "-"
    erro_crit = "SIM" if (raw.get("erro_critico") or c.get("erro_critico") or len(raw.get("erros_fatais_identificados") or c.get("erros_fatais") or []) > 0) else "Não"
    
    transcricao = (
        c.get("transcricao_diarizada") 
        or c.get("diarized_transcript") 
        or c.get("transcricao") 
        or "-"
    )
    
    call_id = c.get("id") or c.get("call_id") or ""
    created_at = c.get("created_at") or c.get("uploaded_at") or c.get("data") or ""
    filename = c.get("filename") or c.get("arquivo") or ""
    setor = raw.get("setor") or c.get("setor") or c.get("equipe") or "Geral"
    
    return {
        "call_id": call_id,
        "created_at": created_at,
        "filename": filename,
        "atendente": atendente,
        "setor": setor,
        "nota_qa": nota_qa,
        "nota_op": nota_op,
        "nps_cli": nps_cli,
        "motivo": motivo,
        "classif": classif,
        "humor_cli": humor_cli,
        "humor_op": humor_op,
        "erro_crit": erro_crit,
        "fase1_qa": fase1_qa,
        "fase1_nps": fase1_nps,
        "fase1_analise": fase1_analise,
        "fase2_qa": fase2_qa,
        "fase2_nps": fase2_nps,
        "fase2_analise": fase2_analise,
        "fase3_qa": fase3_qa,
        "fase3_nps": fase3_nps,
        "fase3_analise": fase3_analise,
        "pontos_positivos": p_pos,
        "pontos_melhoria": p_neg,
        "pontos_positivos_list": p_pos_list if isinstance(p_pos_list, list) else [],
        "pontos_melhoria_list": p_neg_list if isinstance(p_neg_list, list) else [],
        "rec_trein": rec_trein,
        "transcricao": transcricao,
        "status": c.get("status", "")
    }

def generate_excel_report(calls: list[dict]) -> bytes:
    """
    Gera uma planilha Excel analítica profissional contendo:
    1. Resumo Executivo (KPIs & Estatísticas)
    2. Detalhamento Analítico por Chamada (populado com todos os campos)
    3. Transcrições Diarizadas
    """
    wb = Workbook()
    
    # ----------------------------------------------------
    # Estilos Visuais Padronizados (Coherence Theme)
    # ----------------------------------------------------
    title_font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    regular_font = Font(name="Segoe UI", size=10, color="334155")
    
    header_fill = PatternFill(start_color="1A6B52", end_color="1A6B52", fill_type="solid") # Verde Jade Coherence
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Dark slate
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="E2E8F0")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    extracted_calls = [_extract_call_fields(c) for c in calls]
    
    # ----------------------------------------------------
    # TAB 1: Resumo Executivo
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Resumo Executivo"
    
    ws_summary.merge_cells("A1:E1")
    cell_title = ws_summary["A1"]
    cell_title.value = "COHERENCE AI - RELATÓRIO ANALÍTICO DE MONITORIA DE CHAMADAS"
    cell_title.font = title_font
    cell_title.fill = title_fill
    cell_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_summary.row_dimensions[1].height = 40
    
    total_chamadas = len(extracted_calls)
    concluidas = [c for c in extracted_calls if str(c.get("status")).lower().startswith("conclu") or str(c.get("status")).lower() == "finalizado"]
    total_concluidas = len(concluidas)
    
    notas_qa = [float(c["nota_qa"]) for c in concluidas if isinstance(c["nota_qa"], (int, float)) or (isinstance(c["nota_qa"], str) and c["nota_qa"].replace('.', '', 1).isdigit())]
    qa_medio = round(sum(notas_qa) / len(notas_qa), 1) if notas_qa else 0.0
    
    nps_list = []
    for c in concluidas:
        val = c.get("nps_cli")
        if val is not None and val != "-":
            try:
                nps_list.append(float(val))
            except (ValueError, TypeError):
                pass
    nps_medio = round(sum(nps_list) / len(nps_list), 1) if nps_list else 0.0
    
    erros_criticos = sum(1 for c in concluidas if c.get("erro_crit") == "SIM")
    
    kpis = [
        ("Métrica Executiva", "Valor"),
        ("Total de Chamadas Recebidas", total_chamadas),
        ("Chamadas Auditadas com Sucesso", total_concluidas),
        ("QA Score Médio (/100)", f"{qa_medio} / 100"),
        ("NPS Médio de Sentimento (/10)", f"{nps_medio} / 10"),
        ("Alertas de Erros Críticos", erros_criticos),
    ]
    
    ws_summary.append([]) # Linha em branco
    for row_idx, (label, val) in enumerate(kpis, start=3):
        ws_summary.cell(row=row_idx, column=1, value=label)
        ws_summary.cell(row=row_idx, column=2, value=str(val))
        
        c1 = ws_summary.cell(row=row_idx, column=1)
        c2 = ws_summary.cell(row=row_idx, column=2)
        
        if row_idx == 3:
            c1.font = header_font
            c1.fill = header_fill
            c2.font = header_font
            c2.fill = header_fill
        else:
            c1.font = bold_font
            c2.font = regular_font
            c1.border = thin_border
            c2.border = thin_border
            
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 25

    # ----------------------------------------------------
    # TAB 2: Detalhamento Analítico
    # ----------------------------------------------------
    ws_detail = wb.create_sheet(title="Detalhamento Analítico")
    
    headers_detail = [
        "ID Chamada", "Data / Horário", "Arquivo", "Atendente", "Setor",
        "Nota QA Geral", "Nota Operador", "NPS Cliente", "Motivo do Contato",
        "Classificação", "Humor Cliente", "Humor Atendente", "Erro Crítico",
        "Fase 1: QA", "Fase 1: NPS", "Fase 1: Análise",
        "Fase 2: QA", "Fase 2: NPS", "Fase 2: Análise",
        "Fase 3: QA", "Fase 3: NPS", "Fase 3: Análise",
        "Pontos Positivos", "Pontos de Melhoria", "Recomendação de Treinamento"
    ]
    
    ws_detail.append(headers_detail)
    ws_detail.row_dimensions[1].height = 28
    
    for col_num, header in enumerate(headers_detail, 1):
        cell = ws_detail.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    for row_idx, c in enumerate(extracted_calls, start=2):
        row_data = [
            str(c["call_id"]),
            str(c["created_at"]),
            str(c["filename"]),
            str(c["atendente"]),
            str(c["setor"]),
            c["nota_qa"],
            c["nota_op"],
            c["nps_cli"],
            str(c["motivo"]),
            str(c["classif"]),
            str(c["humor_cli"]),
            str(c["humor_op"]),
            c["erro_crit"],
            c["fase1_qa"], c["fase1_nps"], str(c["fase1_analise"]),
            c["fase2_qa"], c["fase2_nps"], str(c["fase2_analise"]),
            c["fase3_qa"], c["fase3_nps"], str(c["fase3_analise"]),
            str(c["pontos_positivos"]),
            str(c["pontos_melhoria"]),
            str(c["rec_trein"])
        ]
        
        ws_detail.append(row_data)
        current_row = ws_detail.row_dimensions[row_idx]
        current_row.height = 45
        
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        
        for col_idx in range(1, len(row_data) + 1):
            cell = ws_detail.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = align_left if col_idx not in [1, 2, 6, 7, 8, 13, 14, 15, 17, 18, 20, 21] else align_center

    # Auto-fit colunas tab 2
    for col in ws_detail.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws_detail.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 50)

    # ----------------------------------------------------
    # TAB 3: Transcrições Diarizadas
    # ----------------------------------------------------
    ws_trans = wb.create_sheet(title="Transcrições Diarizadas")
    ws_trans.append(["ID Chamada", "Arquivo", "Atendente", "Transcrição Diarizada Completa"])
    ws_trans.row_dimensions[1].height = 25
    
    for col_num in range(1, 5):
        cell = ws_trans.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        
    for row_idx, c in enumerate(extracted_calls, start=2):
        row_data = [
            str(c["call_id"]),
            str(c["filename"]),
            str(c["atendente"]),
            str(c["transcricao"])
        ]
        ws_trans.append(row_data)
        ws_trans.row_dimensions[row_idx].height = 60
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        
        for col_idx in range(1, 5):
            cell = ws_trans.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = align_left
            
    ws_trans.column_dimensions['A'].width = 25
    ws_trans.column_dimensions['B'].width = 30
    ws_trans.column_dimensions['C'].width = 20
    ws_trans.column_dimensions['D'].width = 85
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
