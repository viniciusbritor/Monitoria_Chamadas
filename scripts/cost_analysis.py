"""
Planilha de Custos — Multi-Cenarios (500/1000/5000 calls/dia) + Chatbots + Folga 25%
Inclui rateio de custos de desenvolvimento + modelos Bayesianos.
"""
import json, os, random, math
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# ─── AUTH ───
TOKEN_PATH = os.path.expanduser(r"~\.gemini\config\skills\google_calendar_manager\resources\token_drive.json")
OUTFILE = fr"C:\Users\vinic\workspace_antigravity\Monitoria_Chamadas\docs\Custos_Projecao_Completa.xlsx"
FOLDER_ID = "1aNCHHOiQQzquuxzaeQQa8qr3ciZcsfMt"

# ============================================================
# PARAMETERS
# ============================================================
SLACK = 1.25  # 25% folga/seguranca

# Development costs (rateio sobre 12 meses)
DEV_COST_HOURS = 120    # horas de desenvolvimento investidas
DEV_RATE_HOUR = 50      # USD/hora (custo medio de engenharia + cloud dev)
DEV_TOTAL_COST = DEV_COST_HOURS * DEV_RATE_HOUR  # $6,000
DEV_AMORTIZATION_MONTHS = 12
DEV_MONTHLY = DEV_TOTAL_COST / DEV_AMORTIZATION_MONTHS  # $500/mes de rateio

AUDIO_MINUTES = 5
VCPU_HOUR = 0.024
GB_HOUR_MEM = 0.0025
WORKER_CPU = 4
WORKER_MEM = 4
API_CPU = 4
API_MEM = 8

WORKER_COST_HOUR = (WORKER_CPU * VCPU_HOUR) + (WORKER_MEM * GB_HOUR_MEM)  # $0.106/hr
API_COST_HOUR = (API_CPU * VCPU_HOUR) + (API_MEM * GB_HOUR_MEM)  # $0.116/hr

WHISPER_MULT = 0.1
LLM_SEC_PER_CALL = 3
PROCESSING_MIN_PER_CALL = (AUDIO_MINUTES * WHISPER_MULT) + (LLM_SEC_PER_CALL / 60)
CONCURRENCY = 2

TOKENS_IN = 2500
TOKENS_OUT = 1200
DS_INPUT_COST_1M = 0.14
DS_OUTPUT_COST_1M = 0.28
FALLBACK_PCT = 0.05
MM_COST_PER_CALL = 0.001

# Fixed infra
FIRESTORE_COST = 3 * SLACK
STORAGE_COST = 5 * SLACK
PUBSUB_COST = 2 * SLACK
SECRETS_COST = 1 * SLACK
CLOUD_BUILD_COST = 3 * SLACK
REGISTRY_COST = 1 * SLACK

# Chatbot costs (WhatsApp Agent, 5 bots, Cloud Run scenario B)
CHATBOT_CLOUDRUN = 30 * SLACK      # 5x Cloud Run (1vCPU, 1GiB)
CHATBOT_POSTGRES = 10 * SLACK      # Cloud SQL db-f1-micro
CHATBOT_EVOLUTION = 40 * SLACK     # Evolution API Cloud Run
CHATBOT_FIRESTORE = 3 * SLACK      # Firestore extra
CHATBOT_LLM_MONTHLY = 8 * SLACK    # LLM para chatbots (200 msg/dia x 5 bots)

CHATBOT_TOTAL = CHATBOT_CLOUDRUN + CHATBOT_POSTGRES + CHATBOT_EVOLUTION + CHATBOT_FIRESTORE + CHATBOT_LLM_MONTHLY

# Market benchmarks
BENCHMARKS = {
    "CallMiner (enterprise)": (0.05, 0.15),
    "Observe.AI": (0.10, 0.20),
    "Gong.io (sales)": (0.08, 0.12),
    "Chorus.ai": (0.06, 0.10),
    "Cogito": (0.04, 0.08),
}

# Scenarios
SCENARIOS = [500, 1000, 5000]

# ============================================================
# BAYESIAN MONTE CARLO
# ============================================================
def monte_carlo(calls_month, n=10000):
    def whisper_prior():
        return max(0.3, random.lognormvariate(math.log(0.55), 0.2))

    def llm_tokens_prior():
        return (random.randint(2000, 3000), random.randint(800, 1800))

    def conc_prior():
        return random.choice([1, 2, 2, 2, 3])

    costs_t, costs_w, costs_l = [], [], []

    for _ in range(n):
        wm = whisper_prior()
        ti, to = llm_tokens_prior()
        conc = conc_prior()
        proc_min = (AUDIO_MINUTES * wm) + (LLM_SEC_PER_CALL / 60)
        hours_day = (calls_month / 30) / (60 / proc_min * conc)
        worker_var = hours_day * 30 * WORKER_COST_HOUR
        llm_c = (ti * calls_month / 1_000_000 * DS_INPUT_COST_1M) + \
                (to * calls_month / 1_000_000 * DS_OUTPUT_COST_1M)
        total = worker_var + llm_c + FIRESTORE_COST + STORAGE_COST + PUBSUB_COST + SECRETS_COST + CLOUD_BUILD_COST + REGISTRY_COST
        costs_t.append(total)
        costs_w.append(worker_var)
        costs_l.append(llm_c)

    s = sorted(costs_t)
    def ci(d, p): return d[int(p * len(d))]
    return {
        "p10": ci(s, 0.10), "p25": ci(s, 0.25), "p50": s[len(s)//2],
        "p75": ci(s, 0.75), "p90": ci(s, 0.90), "mean": sum(costs_t)/n,
        "mean_worker": sum(costs_w)/n, "mean_llm": sum(costs_l)/n,
    }

# ============================================================
# COST CALCULATOR (deterministic + slack)
# ============================================================
def calc_scenario(calls_day, with_dev=True, with_chatbot=True):
    """Returns dict with all cost components for a given call volume."""
    calls_month = calls_day * 30

    # Worker compute
    calls_hour = (60 / PROCESSING_MIN_PER_CALL) * CONCURRENCY
    var_hours_month = (calls_month / calls_hour)
    worker_var = var_hours_month * WORKER_COST_HOUR
    worker_fixed = WORKER_COST_HOUR * 730  # min=1 always-on

    # LLM
    llm_ds = (TOKENS_IN * calls_month / 1_000_000 * DS_INPUT_COST_1M) + \
             (TOKENS_OUT * calls_month / 1_000_000 * DS_OUTPUT_COST_1M)
    llm_mm = FALLBACK_PCT * calls_month * MM_COST_PER_CALL
    llm_total = llm_ds + llm_mm

    # API
    api_hours = calls_month * 0.01  # ~36s processing per upload
    api_cost = max(1, api_hours * API_COST_HOUR)

    # Storage scales with volume
    storage = max(1, (calls_day / 500) * STORAGE_COST)
    pubsub = max(1, (calls_day / 500) * PUBSUB_COST)
    firestore = max(1, (calls_day / 500) * FIRESTORE_COST)

    # Base costs without slack
    infra_base = firestore + storage + pubsub + SECRETS_COST + CLOUD_BUILD_COST + REGISTRY_COST

    # With slack
    infra = infra_base * SLACK
    worker_var_s = worker_var * SLACK
    worker_fixed_s = worker_fixed * SLACK
    llm_s = llm_total * SLACK
    api_s = api_cost * SLACK

    push_total = worker_var_s + llm_s + api_s + infra
    pull_total = worker_fixed_s + worker_var_s + llm_s + api_s + infra

    # Dev amortization
    dev_monthly = DEV_MONTHLY if with_dev else 0

    # Chatbot
    chatbot = CHATBOT_TOTAL if with_chatbot else 0

    push_final = push_total + dev_monthly
    pull_final = pull_total + dev_monthly

    mono_total = push_final + chatbot  # monitoria + chatbot (PUSH)
    mono_pull_total = pull_final + chatbot

    return {
        "calls_day": calls_day,
        "calls_month": calls_month,
        "worker_var": worker_var_s,
        "worker_fixed": worker_fixed_s,
        "llm_ds": llm_ds * SLACK,
        "llm_mm": llm_mm * SLACK,
        "llm_total": llm_s,
        "api_cost": api_s,
        "infra": infra,
        "push_total": push_total,
        "pull_total": pull_total,
        "dev_rateio": dev_monthly,
        "chatbot_total": chatbot,
        "monitoria_push": push_final,
        "monitoria_pull": pull_final,
        "total_push": mono_total,
        "total_pull": mono_pull_total,
        "per_call_push": push_final / calls_month,
        "per_call_pull": pull_final / calls_month,
    }

# ============================================================
# RUN ALL
# ============================================================
print("Computing scenarios...")
bayes_results = {}
scenario_results = {}

for c in SCENARIOS:
    cm = c * 30
    print(f"  Bayesian {c} calls/day ({cm}/mo)...")
    bayes_results[c] = monte_carlo(cm)
    print(f"  Deterministic {c} calls/day...")
    scenario_results[c] = calc_scenario(c)

# ============================================================
# CREATE EXCEL
# ============================================================
print("Building Excel...")
wb = openpyxl.Workbook()

# Styles
hdr_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
orange_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
usd_fmt = '$#,##0.00'
pct_fmt = '0.0%'
num6_fmt = '$#,##0.000000'
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_row(ws, row, cols, fill=None, bold=False, fmt=usd_fmt):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if fill: cell.fill = fill
        if bold: cell.font = Font(bold=True, size=11)

def hdr(ws, row, headers):
    for i, h in enumerate(headers, 1):
        ws.cell(row=row, column=i, value=h)
        ws.cell(row=row, column=i).font = hdr_font
        ws.cell(row=row, column=i).fill = hdr_fill
        ws.cell(row=row, column=i).alignment = center
        ws.cell(row=row, column=i).border = thin_border

def val(ws, row, col, v, fmt=usd_fmt):
    if isinstance(v, (int, float)):
        v = round(v, 2)
    c = ws.cell(row=row, column=col, value=v)
    if fmt: c.number_format = fmt
    c.border = thin_border
    c.alignment = center
    return c

# ══════ SHEET 1: RESUMO EXECUTIVO ══════
ws1 = wb.active
ws1.title = "Resumo Executivo"
ws1.sheet_properties.tabColor = "1F4E79"

ws1.merge_cells('A1:H1')
ws1.cell(row=1, column=1, value="Resumo Executivo — Projecao de Custos OmniChannel").font = Font(bold=True, size=14, color='1F4E79')
ws1.merge_cells('A2:H2')
ws1.cell(row=2, column=1, value=f"Gerado: {datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M UTC')} | Folga: +25% | Rateio Dev: ${DEV_MONTHLY:.0f}/mes (12 meses) | Cambio: 1 USD = 5.5 BRL").font = Font(color='666666', size=9)

# Key metrics
ws1.cell(row=4, column=1, value="CUSTOS TOTAIS MENSAIS (Monitoria + 5 Chatbots)").font = Font(bold=True, size=12, color='1F4E79')
ws1.merge_cells('A4:H4')

key_h = ["Cenario", "Monitoria (PUSH, min=0)", "Monitoria (PULL, min=1)", "5 Chatbots", "Rateio Dev", "TOTAL (PUSH)", "TOTAL (PULL)", "Custo/Chamada (PUSH)"]
hdr(ws1, 5, key_h)

for r, c in enumerate(SCENARIOS, 6):
    s = scenario_results[c]
    cell = val(ws1, r, 1, f"{c} chamadas/dia ({s['calls_month']:,}/mes)", None)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.font = Font(bold=True)
    val(ws1, r, 2, s['push_total'])
    val(ws1, r, 3, s['pull_total'])
    val(ws1, r, 4, s['chatbot_total'])
    val(ws1, r, 5, s['dev_rateio'])
    val(ws1, r, 6, s['total_push']).font = Font(bold=True, color='006100' if c <= 1000 else 'BF8F00')
    val(ws1, r, 7, s['total_pull']).font = Font(bold=True, color='9C0006')
    val(ws1, r, 8, s['per_call_push'], num6_fmt).font = Font(color='006100')
    set_row(ws1, r, 8, gray_fill if c == 500 else None, c == 500)

# BRL conversion
ws1.cell(row=11, column=1, value=" ").font = Font(size=8)
ws1.cell(row=12, column=1, value="CONVERSAO BRL (1 USD = 5.5 BRL)").font = Font(bold=True, size=12, color='1F4E79')
ws1.merge_cells('A12:H12')

brl_h = ["Cenario", "Monitoria BRL", "Chatbots BRL", "TOTAL BRL/mes", "TOTAL BRL/ano", "", "", ""]
hdr(ws1, 13, brl_h)
for r, c in enumerate(SCENARIOS, 14):
    s = scenario_results[c]
    val(ws1, r, 1, f"{c} calls/dia", None).font = Font(bold=True)
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
    mono_brl = s['monitoria_push'] * 5.5
    chat_brl = s['chatbot_total'] * 5.5
    total_brl = s['total_push'] * 5.5
    val(ws1, r, 2, mono_brl)
    val(ws1, r, 3, chat_brl)
    val(ws1, r, 4, total_brl).font = Font(bold=True, color='006100')
    val(ws1, r, 5, total_brl * 12).font = Font(bold=True, color='1F4E79')
    set_row(ws1, r, 8, gray_fill if c == 500 else None, c == 500)

# Comparison with market
ws1.cell(row=19, column=1, value=" ").font = Font(size=8)
ws1.cell(row=20, column=1, value="COMPARATIVO MERCADO (500 chamadas/dia)").font = Font(bold=True, size=12, color='1F4E79')
ws1.merge_cells('A20:H20')

mkt_h = ["Solucao", "Custo/Chamada", "Custo/Mes (15K calls)", "Economia vs Nos (PUSH)", "", "", "", ""]
hdr(ws1, 21, mkt_h)

for r, (name, (lo, hi)) in enumerate(BENCHMARKS.items(), 22):
    mid = (lo + hi) / 2
    monthly = mid * 15000
    vs_us = monthly - scenario_results[500]['monitoria_push']
    val(ws1, r, 1, name, None).alignment = Alignment(horizontal='left', vertical='center')
    val(ws1, r, 2, mid, num6_fmt)
    val(ws1, r, 3, monthly)
    val(ws1, r, 4, vs_us).font = Font(bold=True, color='006100')
    set_row(ws1, r, 8)

# Nossa linha
r = 22 + len(BENCHMARKS)
val(ws1, r, 1, "NOSSA (PUSH, min=0, +folga 25%)", None).font = Font(bold=True, color='006100')
    ws1.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
val(ws1, r, 2, scenario_results[500]['per_call_push'], num6_fmt).font = Font(bold=True, color='006100')
val(ws1, r, 3, scenario_results[500]['monitoria_push']).font = Font(bold=True, color='006100')
val(ws1, r, 4, 0)
set_row(ws1, r, 8, green_fill, True)

ws1.column_dimensions['A'].width = 30
for c in 'BCDEFGH': ws1.column_dimensions[c].width = 18

# ══════ SHEET 2: DETALHE POR CENARIO ══════
ws2 = wb.create_sheet("Detalhe Cenarios")
ws2.sheet_properties.tabColor = "4472C4"

ws2.merge_cells('A1:G1')
ws2.cell(row=1, column=1, value="Detalhamento de Custos por Componente e Cenario (+25% folga)").font = Font(bold=True, size=14)

detail_headers = ["Componente", "500/dia", "1000/dia", "5000/dia", "Nota", "", ""]
hdr(ws2, 3, detail_headers)

detail_rows = [
    ("Worker (variavel, min=0)",  "worker_var",  "Processamento sob demanda"),
    ("Worker (fixo, min=1)",      "worker_fixed","730h/mes x $0.106/h"),
    ("DeepSeek V4 Flash (LLM)",   "llm_ds",      "Tokens: 2500 in + 1200 out"),
    ("MiniMax M3 (fallback 5%)",  "llm_mm",      "Fallback qdo DeepSeek falha"),
    ("API Cloud Run (4vCPU/8GB)", "api_cost",     "Uso leve, min=0"),
    ("Firestore + Storage + PubSub","infra",      "Scale com volume"),
    ("Rateio Desenvolvimento",    "dev_rateio",   f"${DEV_TOTAL_COST:,.0f} ÷ {DEV_AMORTIZATION_MONTHS} meses"),
    ("5 Chatbots (WhatsApp)",     "chatbot_total","Cloud Run + Postgres + LLM"),
    ("TOTAL (PUSH)",              "total_push",   "Monitoria + Dev + Chatbots"),
    ("TOTAL (PULL)",              "total_pull",   "Monitoria + Dev + Chatbots"),
    ("Custo por Chamada (PUSH)",  "per_call_push","USD/chamada"),
]

for r, (label, key, note) in enumerate(detail_rows, 4):
    ws2.cell(row=r, column=1, value=label).alignment = Alignment(horizontal='left', vertical='center')
    ws2.cell(row=r, column=1).font = Font(bold=True)
    for ci, c in enumerate(SCENARIOS, 2):
        v = scenario_results[c][key]
        if "per_call" in key:
            val(ws2, r, ci, v, num6_fmt)
        else:
            val(ws2, r, ci, v)
    ws2.cell(row=r, column=5, value=note).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws2.cell(row=r, column=5).font = Font(color='666666', size=9)
    is_total = "TOTAL" in label
    set_row(ws2, r, 7, green_fill if "PUSH" in label and "TOTAL" in label else (orange_fill if "PULL" in label and "TOTAL" in label else gray_fill), is_total)

ws2.column_dimensions['A'].width = 32
for c in 'BCDEFG': ws2.column_dimensions[c].width = 16
ws2.column_dimensions['E'].width = 35

# ══════ SHEET 3: BAYESIAN ══════
ws3 = wb.create_sheet("Bayesian")
ws3.sheet_properties.tabColor = "548235"

ws3.merge_cells('A1:F1')
ws3.cell(row=1, column=1, value="Modelo Bayesiano — Monte Carlo (10.000 simulacoes) — com Folga +25%").font = Font(bold=True, size=14)
ws3.merge_cells('A2:F2')
ws3.cell(row=2, column=1, value="Priors: Whisper ~ LogNormal(0.55, 0.2) | Tokens ~ Uniform | Concurrency ~[1,2,2,2,3] | Folga: 1.25x aplicada apos simulacao").font = Font(color='666666', size=9)

b_h = ["Cenario", "Mediana (P50)", "P10 (otimista)", "P25", "P75", "P90 (pessimista)"]
hdr(ws3, 4, b_h)

bf = "$#,##0.00"

for r, c in enumerate(SCENARIOS, 5):
    b = bayes_results[c]
    cell = val(ws3, r, 1, f"{c} chamadas/dia ({c*30:,}/mes)", None)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.font = Font(bold=True)
    vals_sl = [b['p50']*SLACK, b['p10']*SLACK, b['p25']*SLACK, b['p75']*SLACK, b['p90']*SLACK]
    for ci, v in enumerate(vals_sl, 2):
        val(ws3, r, ci, v, bf)
    set_row(ws3, r, 6, gray_fill if c == 500 else None, c == 500)

# Credible interval summary
ws3.cell(row=10, column=1, value="Intervalo de Credibilidade (com folga 25%)").font = Font(bold=True, size=11, color='1F4E79')
ws3.merge_cells('A10:F10')

for r, c in enumerate(SCENARIOS, 11):
    b = bayes_results[c]
    lo = b['p10'] * SLACK
    hi = b['p90'] * SLACK
    ws3.cell(row=r, column=1, value=f"{c} calls/dia:").font = Font(bold=True)
    ws3.merge_cells(f'B{r}:F{r}')
    ws3.cell(row=r, column=2, value=f"${lo:,.2f} — ${hi:,.2f}/mes  (80% confianca)  |  por chamada: ${lo/(c*30):.4f} — ${hi/(c*30):.4f}")
    ws3.cell(row=r, column=2).font = Font(color='006100')

# Per-call Bayesian
ws3.cell(row=15, column=1, value="Custo por Chamada (Mediana, com folga)").font = Font(bold=True, size=11, color='1F4E79')
ws3.merge_cells('A15:F15')

pch = ["Cenario", "P10/chamada", "P50/chamada", "P90/chamada", "", ""]
hdr(ws3, 16, pch)
for r, c in enumerate(SCENARIOS, 17):
    b = bayes_results[c]
    cm = c * 30
    cell = val(ws3, r, 1, f"{c} calls/dia", None)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    val(ws3, r, 2, b['p10']/cm*SLACK, num6_fmt)
    val(ws3, r, 3, b['p50']/cm*SLACK, num6_fmt).font = Font(bold=True, color='006100')
    val(ws3, r, 4, b['p90']/cm*SLACK, num6_fmt)
    set_row(ws3, r, 6)

ws3.column_dimensions['A'].width = 22
for c in 'BCDEF': ws3.column_dimensions[c].width = 18

# ══════ SHEET 4: PREMISSAS ══════
ws4 = wb.create_sheet("Premissas")
ws4.sheet_properties.tabColor = "BF8F00"

ws4.merge_cells('A1:D1')
ws4.cell(row=1, column=1, value="Premissas e Fontes do Calculo").font = Font(bold=True, size=14)

prem_h = ["Parametro", "Valor", "Fonte", "Observacao"]
hdr(ws4, 3, prem_h)

premissas = [
    ("Chamadas/Dia (cenarios)", "500 / 1.000 / 5.000", "Usuario", ""),
    ("Chamadas/Mes", "15.000 / 30.000 / 150.000", "Calculado", "Calls × 30"),
    ("Audio medio", "5 minutos", "Estimado", ""),
    ("Modelo Whisper", "base (74MB, int8)", "faster-whisper", "~0.1x real-time"),
    ("Concorrencia worker", "2", "cloudbuild-worker.yaml", ""),
    ("Worker vCPU / RAM", "4 vCPU / 4 GiB", "cloudbuild-worker.yaml", ""),
    ("Cloud Run vCPU-hr", "$0.024/hr", "GCP Pricing jul/2026", ""),
    ("DeepSeek V4 Flash input", "$0.14/1M tokens", "api-docs.deepseek.com", ""),
    ("DeepSeek V4 Flash output", "$0.28/1M tokens", "api-docs.deepseek.com", ""),
    ("Tokens input/call (avg)", "2.500", "Estimado", "Prompt + contexto"),
    ("Tokens output/call (avg)", "1.200", "Estimado", "JSON resposta"),
    ("Fallback MiniMax %", "5%", "Estimado", "Taxa de falha DeepSeek"),
    ("Folga de Seguranca", "+25%", "Usuario", "Margem de seguranca"),
    ("Horas Desenvolvimento", "120 horas", "Estimado", ""),
    ("Custo Hora Dev", "$50/h", "Mercado BR", "Engenharia + cloud dev"),
    ("Custo Total Dev", f"${DEV_TOTAL_COST:,.0f}", "Calculado", "120h × $50/h"),
    ("Rateio Dev (meses)", "12", "Configuracao", f"${DEV_MONTHLY:,.0f}/mes"),
    ("Chatbots (5 simultaneos)", "Cloud Run + Postgres", "Cenario B (docs/CUSTOS.md)", f"${CHATBOT_TOTAL:,.0f}/mes"),
    ("Simulacoes Monte Carlo", "10.000", "Configuracao", ""),
    ("Cambio USD/BRL", "5.5", "Mercado jul/2026", ""),
    ("Projeto GCP Principal", "coherence-ominichannel-fs", "Cloud Run/Firestore/PubSub", ""),
    ("Projeto GCP WhatsApp", "jennifer-bot", "Compute Engine/Firestore", ""),
]

for r, (param, val, source, obs) in enumerate(premissas, 4):
    ws4.cell(row=r, column=1, value=param).alignment = Alignment(horizontal='left', vertical='center')
    ws4.cell(row=r, column=2, value=val).alignment = center
    ws4.cell(row=r, column=3, value=source).alignment = Alignment(horizontal='left', vertical='center')
    ws4.cell(row=r, column=4, value=obs).alignment = Alignment(horizontal='left', vertical='center')
    set_row(ws4, r, 4, gray_fill if r % 2 == 0 else None)

ws4.column_dimensions['A'].width = 30
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 25
ws4.column_dimensions['D'].width = 30

# ══════ SHEET 5: CHATBOTS ══════
ws5 = wb.create_sheet("Chatbots")
ws5.sheet_properties.tabColor = "7030A0"

ws5.merge_cells('A1:E1')
ws5.cell(row=1, column=1, value="Infraestrutura WhatsApp — 5 Chatbots Simultaneos (com folga 25%)").font = Font(bold=True, size=14)

cb_h = ["Componente", "Configuracao", "Custo Base/mes", "Custo +25% Folga", "Nota"]
hdr(ws5, 3, cb_h)

chatbot_items = [
    ("5x Cloud Run (agentes)", "1 vCPU, 1 GiB, min=0", 30, CHATBOT_CLOUDRUN, "Sob demanda, escala a zero"),
    ("Cloud SQL Postgres", "db-f1-micro, 10 GB", 10, CHATBOT_POSTGRES, "Banco compartilhado"),
    ("Evolution API (Cloud Run)", "1 vCPU, 512 MiB, min=0", 40, CHATBOT_EVOLUTION, "Gerenciamento WhatsApp"),
    ("Firestore extra", "Read/write 5x", 3, CHATBOT_FIRESTORE, "Sessoes, personas, conhecimento"),
    ("LLM (DeepSeek Flash)", "~200 msg/dia x 5 bots", 8, CHATBOT_LLM_MONTHLY, "~1000 tokens/msg"),
]

for r, (comp, config, base, with_folga, note) in enumerate(chatbot_items, 4):
    ws5.cell(row=r, column=1, value=comp).alignment = Alignment(horizontal='left', vertical='center')
    ws5.cell(row=r, column=2, value=config).alignment = center
    val(ws5, r, 3, base)
    val(ws5, r, 4, with_folga)
    ws5.cell(row=r, column=5, value=note).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    set_row(ws5, r, 5, gray_fill if r % 2 == 0 else None)

r += 1
val(ws5, r, 1, "TOTAL 5 Chatbots", None).font = Font(bold=True, size=12)
ws5.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
val(ws5, r, 4, CHATBOT_TOTAL).font = Font(bold=True, color='7030A0', size=12)
set_row(ws5, r, 5, green_fill, True)

ws5.column_dimensions['A'].width = 30
ws5.column_dimensions['B'].width = 28
ws5.column_dimensions['C'].width = 18
ws5.column_dimensions['D'].width = 18
ws5.column_dimensions['E'].width = 35

# ══════ SAVE & UPLOAD ══════
wb.save(OUTFILE)
print(f"Saved: {OUTFILE}")

# Upload to Drive
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

with open(TOKEN_PATH) as f:
    creds_data = json.load(f)
creds = Credentials.from_authorized_user_info(creds_data, scopes=["https://www.googleapis.com/auth/drive"])
service = build("drive", "v3", credentials=creds)

# Delete old version if exists
old = service.files().list(q=f"name='Custos_Projecao_Completa.xlsx' and '{FOLDER_ID}' in parents and trashed=false",
                           fields="files(id)").execute()
for f in old.get("files", []):
    service.files().delete(fileId=f["id"]).execute()
    print(f"Deleted old version: {f['id']}")

meta = {"name": "Custos_Projecao_Completa.xlsx", "parents": [FOLDER_ID]}
media = MediaFileUpload(OUTFILE, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
up = service.files().create(body=meta, media_body=media, fields="id, webViewLink").execute()

print(f"\nUpload: {up['webViewLink']}")
print(f"Local: {OUTFILE}")
print(f"Folder: https://drive.google.com/drive/folders/{FOLDER_ID}")

# ─── PRINT SUMMARY ───
print("\n" + "="*70)
print("RESUMO (com folga 25% + rateio dev):")
print("="*70)
for c in SCENARIOS:
    s = scenario_results[c]
    print(f"\n{c} chamadas/dia ({s['calls_month']:,}/mes):")
    print(f"  Monitoria (PUSH):   ${s['monitoria_push']:,.2f}/mes  |  ${s['per_call_push']:.4f}/chamada")
    print(f"  Monitoria (PULL):   ${s['monitoria_pull']:,.2f}/mes")
    print(f"  +5 Chatbots:        ${s['chatbot_total']:,.2f}/mes")
    print(f"  TOTAL (PUSH+Chat):  ${s['total_push']:,.2f}/mes  (BRL: R$ {s['total_push']*5.5:,.2f})")
    print(f"  TOTAL/ano:          ${s['total_push']*12:,.2f}  (BRL: R$ {s['total_push']*12*5.5:,.2f})")
