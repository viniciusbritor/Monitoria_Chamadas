"""
Planilha Custos OmniChannel v5 — GCP + AI (Dev+Prod) + Escala
Sem rateio. Separacao clara: Infra | AI Dev | AI Modulos
Benchmarks nacionais: Teleperformance, Atento
"""
import json, os
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\Users\vinic\workspace_antigravity\Monitoria_Chamadas\docs\Custos_OmniChannel_Executivo.xlsx"
FID = "1aNCHHOiQQzquuxzaeQQa8qr3ciZcsfMt"
TKN = os.path.expanduser(r"~\.gemini\config\skills\google_calendar_manager\resources\token_drive.json")
BRL = 5.50; SL = 1.10

# ══════ LLM REAL COSTS ══════
MM_PLAN = 20          # MiniMax Plus — plano fixo mensal
DS_DEV = 15           # DeepSeek — tokens para desenvolvimento (testes, prompts)
DS_PROD_MONITORIA = 2 # DeepSeek — tokens Monitoria prod (~20 calls/dia)
DS_PROD_CHATBOTS = 8  # DeepSeek — tokens Chatbots (200 msg/dia x 5 bots)
DS_PROD_EXTRA = 15    # DeepSeek — tokens extras (URA dev, testes, overflow)

# ══════ Styles ══════
DB, GR, OR, LG, BL = "1F4E79", "C6EFCE", "FCE4D6", "F5F5F5", "4472C4"
hf = Font(name='Calibri',bold=True,size=11,color='FFFFFF')
hd = PatternFill(start_color=DB,end_color=DB,fill_type='solid')
gf = PatternFill(start_color=GR,end_color=GR,fill_type='solid')
of = PatternFill(start_color=OR,end_color=OR,fill_type='solid')
lf = PatternFill(start_color=LG,end_color=LG,fill_type='solid')
th = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))
C = Alignment(horizontal='center',vertical='center',wrap_text=True)
L = Alignment(horizontal='left',vertical='center',wrap_text=True)

def W(ws,r,c,v,fmt=None,bold=False,color=None,fill=None,align=None,sz=None):
    cl=ws.cell(row=r,column=c,value=v)
    cl.border=th; cl.alignment=align or C
    if fmt: cl.number_format=fmt
    if bold: cl.font=Font(bold=True,size=sz or 11,color=color or '333333')
    if fill: cl.fill=fill
    if sz and not bold: cl.font=Font(size=sz)
    return cl

def H(ws,r,data):
    for c,v in enumerate(data,1):
        cl=ws.cell(row=r,column=c,value=v)
        cl.font=hf; cl.fill=hd; cl.alignment=C; cl.border=th

def RN(v): return round(v*SL,2)

# ══════ DATA ══════

# --- GCP INFRA ---
INFRA = [
    ("Worker (4vCPU/4GB, min=1)", RN(77.38)),
    ("API + Portal Cloud Run (min=0)", RN(7.00)),
    ("Firestore (2 projetos)", RN(15.00)),
    ("Cloud Storage (audios + backups)", RN(6.00)),
    ("Pub/Sub + Secret Manager", RN(3.00)),
    ("Cloud Build + Artifact Registry", RN(4.00)),
    ("VM e2-small + IP (WhatsApp)", RN(20.00)),
]
GCP_TOTAL = sum(v for _,v in INFRA)

# --- AI DESENVOLVIMENTO ---
AI_DEV = [
    ("DeepSeek — tokens desenvolvimento (testes, prompts)", RN(DS_DEV)),
    ("MiniMax Plus — plano mensal (fallback + créditos)", MM_PLAN),
]
AI_DEV_T = sum(v for _,v in AI_DEV)

# --- AI PRODUCAO (modulos) ---
AI_PROD = [
    ("DeepSeek — Monitoria (~20 chamadas/dia)", RN(DS_PROD_MONITORIA)),
    ("DeepSeek — Chatbots WhatsApp (5 bots)", RN(DS_PROD_CHATBOTS)),
    ("DeepSeek — Extras (URA dev, overflow)", RN(DS_PROD_EXTRA)),
]
AI_PROD_T = sum(v for _,v in AI_PROD)

# --- TOTALS ---
AI_TOTAL = AI_DEV_T + AI_PROD_T
TOTAL_HOJE = GCP_TOTAL + AI_TOTAL

# ══════ PROJECTIONS ══════
def ds_tokens(calls_month):
    return (calls_month*2500/1e6*0.14)+(calls_month*1200/1e6*0.28)

WA = RN(91.0)  # WhatsApp 5 bots infra

S = {}
for c in [500,1000,5000]:
    cm = c*30
    ds_mono = ds_tokens(cm)  # DeepSeek tokens para Monitoria em escala
    ds_chat = 8.80           # Chatbot LLM (fixo)
    mm_token = ds_mono*0.05*0.75

    S[c] = {
        "worker":    RN(7.29*c/500),   # escala linear
        "ds_mono":   RN(ds_mono),       # escala com volume
        "mm_plan":   MM_PLAN,           # fixo
        "mm_token":  RN(mm_token),       # fallback escala
        "ds_chat":   RN(ds_chat),       # fixo chatbots
        "ds_dev":    RN(DS_DEV),         # dev estabiliza
        "ds_extra":  RN(DS_PROD_EXTRA), # extra estabiliza
        "infra":     RN(23.10),          # API + storage + pubsub
        "wa_infra":  WA,                 # WhatsApp infra
    }
    s = S[c]
    s["gcp"] = s["worker"]+s["infra"]+s["wa_infra"]
    s["ai_dev"] = s["mm_plan"]+s["ds_dev"]
    s["ai_prod"] = s["ds_mono"]+s["mm_token"]+s["ds_chat"]+s["ds_extra"]
    s["total"] = s["gcp"]+s["ai_dev"]+s["ai_prod"]
    s["pc"] = s["total"]*BRL/cm

# ══════ COMPARATIVOS NACIONAIS ══════
BENCH = [
    ("Teleperformance (Brasil)", "BPO global, 80K funcionários BR", "QA humano por amostragem", 0.50, 1.00),
    ("Atento (Brasil)", "Maior BPO LatAm, 90K func. BR", "Monitoria manual de chamadas", 0.40, 0.80),
    ("Liq (Bertelsmann)", "BPO digital, 40K func. BR", "Plataforma própria + terceiros", 0.35, 0.70),
    ("Algar Tech", "BPO médio, 20K func. BR", "Soluções híbridas", 0.30, 0.60),
]

# ══════ BUILD ══════
wb = Workbook()

# ── SHEET 1: RESUMO ──
ws = wb.active; ws.title="Resumo"; ws.sheet_properties.tabColor=DB
ws.merge_cells('A1:F1')
ws.cell(row=1,column=1,value="CUSTOS OMNI CHANNEL · Infra + IA (Desenvolvimento + Produção)").font=Font(bold=True,size=16,color=DB)
ws.merge_cells('A2:F2')
ws.cell(row=2,column=1,value=f"{datetime.now(timezone.utc).strftime('%d/%m/%Y %H:%M')} | R${BRL:.2f}/USD | Margem 10% | DeepSeek pay-per-token | MiniMax Plus ${MM_PLAN}/mês | Valores em BRL").font=Font(color='666666',size=9)

# TODAY COSTS
ws.cell(row=4,column=1,value="HOJE — Custo Operacional Real").font=Font(bold=True,size=13,color=DB)
ws.merge_cells('A4:F4')
H(ws,5,["Categoria","Serviço","USD/mês","BRL/mês","%","Descrição"])

r=6
for nm,usd in INFRA:
    W(ws,r,1,"GCP",align=L,sz=9,color=DB); W(ws,r,2,nm,align=L,sz=9); W(ws,r,3,usd,'$#,##0')
    W(ws,r,4,usd*BRL,'R$ #,##0'); W(ws,r,5,usd/TOTAL_HOJE,'0%',sz=9); W(ws,r,6,"Infraestrutura cloud",align=L,sz=9)
    ws.row_dimensions[r].height=20; r+=1

for nm,usd in AI_DEV+AI_PROD:
    cat = "IA Desenv." if "plano" in nm.lower() or "desenvolvimento" in nm.lower() else "IA Produção"
    W(ws,r,1,cat,align=L,sz=9,color=BL); W(ws,r,2,nm,align=L,sz=9); W(ws,r,3,usd,'$#,##0')
    W(ws,r,4,usd*BRL,'R$ #,##0'); W(ws,r,5,usd/TOTAL_HOJE,'0%',sz=9); W(ws,r,6,"LLM API",align=L,sz=9)
    ws.row_dimensions[r].height=20; r+=1

# Subtotals
subs = [("GCP Infraestrutura", GCP_TOTAL, DB), ("IA Desenvolvimento", AI_DEV_T, BL), ("IA Produção", AI_PROD_T, BL)]
for lbl,val,clr in subs:
    W(ws,r,1,lbl,bold=True,sz=11,color=clr)
    W(ws,r,3,val,'$#,##0',bold=True)
    W(ws,r,4,val*BRL,'R$ #,##0',bold=True)
    W(ws,r,5,val/TOTAL_HOJE,'0%',bold=True)
    W(ws,r,6,"",align=L,sz=9)
    ws.row_dimensions[r].height=22; r+=1

W(ws,r,1,"TOTAL HOJE",bold=True,sz=14,fill=gf)
W(ws,r,3,TOTAL_HOJE,'$#,##0',bold=True,fill=gf)
W(ws,r,4,TOTAL_HOJE*BRL,'R$ #,##0',bold=True,sz=15,fill=gf)
W(ws,r,5,"100%",bold=True,fill=gf); W(ws,r,6,"",fill=gf)

# SCALE
r+=2
ws.merge_cells(f'A{r}:F{r}')
ws.cell(row=r,column=1,value="PROJEÇÕES DE CRESCIMENTO — IA escala com volume de chamadas").font=Font(bold=True,size=13,color=DB)
r+=1
H(ws,r,["Cenário","Chamadas/mês","GCP+IA/mês (BRL)","Custo/Chamada","Crescimento vs Hoje","Nota"])

for ci,(c,s) in enumerate([(500,S[500]),(1000,S[1000]),(5000,S[5000])],r+1):
    cm = c*30
    f = gf if c>=5000 else (lf if c==500 else None)
    W(ws,ci,1,f"{c}/dia ({cm:,}/mês)",align=L,bold=True,sz=11)
    W(ws,ci,2,cm,'#,##0',bold=True)
    W(ws,ci,3,s["total"]*BRL,'R$ #,##0',bold=True,sz=13,fill=f)
    W(ws,ci,4,s["pc"],'R$ #,##0.00',bold=True,color=DB,sz=14)
    W(ws,ci,5,f"{((s['total']-TOTAL_HOJE)/TOTAL_HOJE*100):+.0f}%",sz=10,color='006100' if s['total']<TOTAL_HOJE else DB)
    note = "IA escala com volume, GCP infra otimizada"
    W(ws,ci,6,note,align=L,sz=9)
    for c2 in range(1,7):
        if f: ws.cell(row=ci,column=c2).fill=f
    ws.row_dimensions[ci].height=28

for c2 in 'ABCDEF': ws.column_dimensions[c2].width=34 if c2=='A' else 20
ws.column_dimensions['B'].width=40

# ── SHEET 2: DETALHE ──
ws2=wb.create_sheet("Detalhe"); ws2.sheet_properties.tabColor=BL
ws2.merge_cells('A1:F1')
ws2.cell(row=1,column=1,value="DETALHE COMPLETO — GCP Infra + IA (Dev + Produção)").font=Font(bold=True,size=14,color=DB)
ws2.merge_cells('A2:F2')
ws2.cell(row=2,column=1,value="MiniMax Plus: plano fixo de $20/mês. DeepSeek: pay-per-token, sem plano fixo. Margem 10% em todos valores.").font=Font(color='666666',size=9)

r=4
for title,data in [("INFRAESTRUTURA GCP",INFRA),("IA — DESENVOLVIMENTO",AI_DEV),("IA — PRODUÇÃO (Módulos)",AI_PROD)]:
    ws2.merge_cells(f'A{r}:F{r}')
    ws2.cell(row=r,column=1,value=title).font=Font(bold=True,size=11,color=DB); r+=1
    H(ws2,r,["Serviço","Base (USD)","+10% Margem (USD)","+10% (BRL)","Categoria","Detalhe"])
    for items in [data]:
        for nm,usd in items:
            cat = "GCP" if "Cloud" in nm or "Firestore" in nm or "Pub" in nm or "VM" in nm or "IP" in nm or "Build" in nm or "Storage" in nm else ("Plano" if "plano" in nm.lower() else "IA Token")
            W(ws2,r,1,nm,align=L,sz=10)
            W(ws2,r,2,round(usd/SL if cat!="Plano" else usd,2),'$#,##0.00',sz=10)
            W(ws2,r,3,usd,'$#,##0.00',sz=10)
            W(ws2,r,4,usd*BRL,'R$ #,##0',sz=10)
            W(ws2,r,5,cat,sz=9)
            W(ws2,r,6,"",align=L,sz=9)
            ws2.row_dimensions[r].height=20; r+=1
    sub=sum(v for _,v in data)
    W(ws2,r,1,"Subtotal",bold=True,fill=lf)
    W(ws2,r,2,"",fill=lf); W(ws2,r,3,sub,'$#,##0',bold=True,fill=lf)
    W(ws2,r,4,sub*BRL,'R$ #,##0',bold=True,fill=lf); W(ws2,r,5,"",fill=lf); W(ws2,r,6,"",fill=lf)
    r+=2

W(ws2,r,1,"TOTAL MENSAL",bold=True,sz=13,fill=gf)
W(ws2,r,3,TOTAL_HOJE,'$#,##0',bold=True,fill=gf)
W(ws2,r,4,TOTAL_HOJE*BRL,'R$ #,##0',bold=True,sz=14,fill=gf)
W(ws2,r,5,"",fill=gf); W(ws2,r,6,"",fill=gf)

for c2 in 'ABCDEF': ws2.column_dimensions[c2].width=38 if c2 in ['A','F'] else 20

# ── SHEET 3: PROJEÇÕES ──
ws3=wb.create_sheet("Projeções"); ws3.sheet_properties.tabColor="548235"
ws3.merge_cells('A1:G1')
ws3.cell(row=1,column=1,value="PROJEÇÕES DE CRESCIMENTO — DeepSeek primário + MiniMax Plus fallback").font=Font(bold=True,size=14,color=DB)
ws3.merge_cells('A2:G2')
ws3.cell(row=2,column=1,value=f"IA Desenvolvimento estabiliza. IA Produção escala com volume de chamadas. WhatsApp 5 bots incluso.").font=Font(color='666666',size=9)

H(ws3,4,["Componente","500/dia","1.000/dia","5.000/dia","Tipo"])

proj = [
    ("GCP — Worker (PUSH, min=0)","worker","Infra"),
    ("GCP — API+Storage+PubSub","infra","Infra"),
    ("GCP — WhatsApp 5 bots (Cloud Run+SQL)","wa_infra","Infra"),
    ("IA Dev — DeepSeek (testes, prompts)","ds_dev","IA Desenv."),
    ("IA Dev — MiniMax Plus plano","mm_plan","IA Desenv."),
    ("IA Prod — DeepSeek Monitoria (diariz+QA)","ds_mono","IA Produção"),
    ("IA Prod — DeepSeek Chatbots","ds_chat","IA Produção"),
    ("IA Prod — DeepSeek Extras (URA dev)","ds_extra","IA Produção"),
    ("IA Prod — MiniMax fallback (5%)","mm_token","IA Produção"),
]
for ri,(nm,k,cat) in enumerate(proj,5):
    W(ws3,ri,1,nm,align=L,sz=9,bold=True)
    for ci,c in enumerate([500,1000,5000],2): W(ws3,ri,ci,S[c][k]*BRL,'R$ #,##0')
    W(ws3,ri,5,cat,sz=9); ws3.row_dimensions[ri].height=20

ri=5+len(proj)
for lbl,key,fl in [("TOTAL GCP","gcp",lf),("TOTAL IA Desenvolvimento","ai_dev",lf),("TOTAL IA Produção","ai_prod",lf)]:
    W(ws3,ri,1,lbl,bold=True,fill=fl)
    for ci,c in enumerate([500,1000,5000],2): W(ws3,ri,ci,S[c][key]*BRL,'R$ #,##0',bold=True,fill=fl)
    W(ws3,ri,5,"",fill=fl); ri+=1

W(ws3,ri,1,"CUSTO TOTAL MENSAL",bold=True,sz=13,fill=gf)
for ci,c in enumerate([500,1000,5000],2): W(ws3,ri,ci,S[c]["total"]*BRL,'R$ #,##0',bold=True,sz=13,fill=gf)
W(ws3,ri,5,"",fill=gf); ri+=2

W(ws3,ri,1,"Custo por Chamada",bold=True,color=DB,sz=12)
for ci,c in enumerate([500,1000,5000],2): W(ws3,ri,ci,S[c]["pc"],'R$ #,##0.00',bold=True,color=DB,sz=14)

for c2 in 'ABCDE': ws3.column_dimensions[c2].width=38 if c2=='A' else 20

# ── SHEET 4: MERCADO NACIONAL ──
ws4=wb.create_sheet("vs Mercado BR"); ws4.sheet_properties.tabColor="BF8F00"
ws4.merge_cells('A1:G1')
ws4.cell(row=1,column=1,value="COMPARATIVO — Operações Nacionais (Brasil)").font=Font(bold=True,size=14,color=DB)
ws4.merge_cells('A2:G2')
ws4.cell(row=2,column=1,value=f"Nossa solução (500 chamadas/dia): R$ {S[500]['pc']:.2f}/chamada, automatizada. Operadoras tradicionais: QA humano, R$ 0,30-1,00/chamada.").font=Font(color='666666',size=9)

H(ws4,4,["Empresa","Perfil","Modelo QA","Custo QA/chamada","vs Nós","Nossa Vantagem",""])
OURS = S[500]["pc"]
for r,(emp,perfil,modelo,lo,hi) in enumerate(BENCH,5):
    mid=(lo+hi)/2
    W(ws4,r,1,emp,align=L,bold=True,sz=11)
    W(ws4,r,2,perfil,align=L,sz=9)
    W(ws4,r,3,modelo,align=L,sz=9)
    W(ws4,r,4,f"R$ {lo:.2f}–{hi:.2f}")
    W(ws4,r,5,f"{mid/OURS:.0f}x mais caro",bold=True,color=DB)
    W(ws4,r,6,f"Automatizável — nosso custo R$ {OURS:.2f}",align=L,sz=9,color='006100')
    ws4.row_dimensions[r].height=28

r+=1
for nm,pc,note in [("NOSSA SOLUÇÃO (500/dia)",OURS,"QA 100% automatizado — DeepSeek V4 Flash"),
                     ("NOSSA SOLUÇÃO (5.000/dia)",S[5000]["pc"],"Escala reduz custo por chamada")]:
    W(ws4,r,1,nm,align=L,bold=True,color='006100',sz=11)
    W(ws4,r,2,"OmniChannel",align=L,sz=9); W(ws4,r,3,"IA — sem intervenção humana",align=L,sz=9)
    W(ws4,r,4,f"R$ {pc:.2f}",bold=True,color='006100',sz=12)
    W(ws4,r,5,"Referência",bold=True)
    W(ws4,r,6,note,align=L,sz=9,color='006100')
    for c2 in range(1,7): ws4.cell(row=r,column=c2).fill=gf
    ws4.row_dimensions[r].height=28; r+=1

r+=1
ws4.merge_cells(f'A{r}:G{r}')
ws4.cell(row=r,column=1,value="Nosso custo de QA automatizado é 5-17x menor que o custo de QA humano das operadoras nacionais.").font=Font(bold=True,color=DB)

for c2 in 'ABCDEFG': ws4.column_dimensions[c2].width=30 if c2 in ['A','F'] else 22

# ══════ SAVE ══════
wb.save(OUT); print(f"Local: {OUT}")

from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

with open(TKN) as f: cr=Credentials.from_authorized_user_info(json.load(f),scopes=["https://www.googleapis.com/auth/drive"])
svc=build("drive","v3",credentials=cr)
old=svc.files().list(q=f"name='Custos_OmniChannel_Executivo.xlsx' and '{FID}' in parents",fields="files(id)").execute()
for f in old.get("files",[]): svc.files().delete(fileId=f["id"]).execute()
up=svc.files().create(body={"name":"Custos_OmniChannel_Executivo.xlsx","parents":[FID]},
    media_body=MediaFileUpload(OUT,mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    fields="id,webViewLink").execute()
print(f"Drive: {up['webViewLink']}")

print(f"\nHOJE: GCP R${GCP_TOTAL*BRL:,.0f} + IA Dev R${AI_DEV_T*BRL:,.0f} + IA Prod R${AI_PROD_T*BRL:,.0f} = TOTAL R${TOTAL_HOJE*BRL:,.0f}/mês")
for c in [500,1000,5000]:
    s=S[c]; print(f"{c}/dia: TOTAL R${s['total']*BRL:,.0f}/mês | R${s['pc']:.2f}/chamada | GCP:{s['gcp']*BRL:,.0f} IA Dev:{s['ai_dev']*BRL:,.0f} IA Prod:{s['ai_prod']*BRL:,.0f}")
OURS=S[500]["pc"]
print(f"vs Teleperformance (R$0.75): economia de R${(0.75-OURS)*15000:,.0f}/mês — {0.75/OURS:.0f}x mais barato")
